"""Build the reviewable workbook: every eligible parcel, re-rankable by weight."""
import copy
import json

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from farmsearch.config import Config
from farmsearch.deliverables import owner_list, rank_shortlist

OUT = "outputs/farm_screen.xlsx"
INK, FIELD, PAPER = "16201A", "3F6F52", "F6F8F4"
BLUE, RULE = "0000FF", "D5DDD4"
HEAD = PatternFill("solid", fgColor=FIELD)
INPUT = PatternFill("solid", fgColor="FFFDE7")
BAND = PatternFill("solid", fgColor="EEF2EC")
THIN = Side(style="thin", color=RULE)
BOX = Border(bottom=THIN)
F = "Arial"


def style_header(ws, row=1, upto=None):
    for c in ws[row]:
        if upto and c.column > upto:
            break
        c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


cfg = Config.load("config/pipeline.yaml")
scored = gpd.read_file("outputs/parcels_scored.gpkg")
summary = json.load(open("outputs/summary.json"))

big = copy.deepcopy(cfg); big.shortlist.top_n = 100000
elig, excluded = rank_shortlist(scored, big)
unit = copy.deepcopy(cfg); unit.shortlist.top_n = 100000
unit.shortlist.weights = {k: 1.0 for k in cfg.shortlist.weights}
norm, _ = rank_shortlist(scored, unit)

METRICS = [c[6:] for c in norm.columns if c.startswith("score_")]
N = norm.set_index("account_id")[[f"score_{m}" for m in METRICS]]
E = elig.set_index("account_id")
P = scored.set_index("account_id")
shipped = dict(zip(pd.read_csv("outputs/shortlist.csv", dtype={"account_id": str})["account_id"],
                   pd.read_csv("outputs/shortlist.csv", dtype={"account_id": str})["rank"]))

wb = Workbook()

# ---------------------------------------------------------------- Read me
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 30, "C": 96})
rows = [
    ("h1", "Farmland screen — Frederick, Carroll and Washington counties", ""),
    ("sub", "Run of 2026-09-03. 129,019 parcels examined, 2,576 scored in full, 1,303 eligible, 40 on the shipped shortlist.", ""),
    ("gap", "", ""),
    ("h2", "The sheets", ""),
    ("kv", "Candidates", "Every parcel that passed the hard rules (1,303). This is the sheet to work in: filter, sort, and record what you decide."),
    ("kv", "Weights", "What each measure is worth. Change a weight and every Score and Rank on Candidates recalculates."),
    ("kv", "Excluded", "The 1,273 parcels a hard rule removed, each with the reason. Nothing here was deleted quietly."),
    ("kv", "Owners", "One row per owner mailbox (2,022). One farmer often holds several parcels; six letters to one address destroys credibility."),
    ("kv", "Comps", "The 102 arms-length agricultural sales behind the per-acre bands, with how each was classified."),
    ("gap", "", ""),
    ("h2", "How to use it", ""),
    ("kv", "Yellow cells", "Yours to edit: the weights, and the Status and Notes columns on Candidates. Everything else is measured or calculated."),
    ("kv", "Score", "Each measure is scaled across the eligible field, then multiplied by its weight and summed. The sign of a weight is its direction: negative means less is better."),
    ("kv", "Rank and Score", "Formulas: your spreadsheet fills them in when it opens the file. 'Score as shipped' is the delivered number, kept beside them so you can always see both."),
    ("kv", "Shipped rank", "Recalculates from Score. 'Shipped rank' is where the parcel sat in the delivered top 40, so you can see what your re-weighting changed."),
    ("gap", "", ""),
    ("h2", "What the numbers do not say", ""),
    ("kv", "Commute", "Free-flow driving times multiplied by a peak factor (BWI 1.35, Langley and Tysons 1.7), not measured traffic. A departure-time-aware routing key would make them real."),
    ("kv", "Approved homes nearby", "Only Frederick publishes approved-but-unbuilt unit counts. Elsewhere the column is empty, which means unknown, not zero."),
    ("kv", "Owner names", "Not public in the Maryland parcel data. The mailing address is the identifier; look the account up on SDAT for the name."),
    ("kv", "Value", "Built from recent arms-length agricultural sales, not from assessments, which understate farmland badly. It is a band, not an appraisal."),
    ("gap", "", ""),
    ("h2", "Six questions no dataset answers", ""),
]
for item in summary.get("cannot_determine", []):
    rows.append(("bullet", "", item))
r = 2
for kind, a, b in rows:
    if kind == "gap":
        r += 1
        continue
    if kind == "h1":
        c = ws.cell(r, 2, a); c.font = Font(name=F, size=16, bold=True, color=INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    elif kind == "sub":
        c = ws.cell(r, 2, a); c.font = Font(name=F, size=10, color="66756C")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    elif kind == "h2":
        c = ws.cell(r, 2, a); c.font = Font(name=F, size=11, bold=True, color=FIELD)
    elif kind == "kv":
        k = ws.cell(r, 2, a); k.font = Font(name=F, size=10, bold=True, color=INK)
        k.alignment = Alignment(vertical="top")
        v = ws.cell(r, 3, b); v.font = Font(name=F, size=10, color=INK)
        v.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
    else:
        v = ws.cell(r, 3, "• " + b); v.font = Font(name=F, size=10, color=INK)
        v.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
    r += 1

# ---------------------------------------------------------------- Weights
wsw = wb.create_sheet("Weights")
wsw.sheet_view.showGridLines = False
widths(wsw, {"A": 42, "B": 11, "C": 13, "D": 62})
for i, h in enumerate(["Measure", "Weight", "Direction", "What it means"], start=1):
    wsw.cell(1, i, h)
style_header(wsw, 1, 4)
MEANING = {
    "largest_contiguous_reachable_acres": "Acres in the biggest block you can drive to from a road entrance.",
    "usable_pct": "Share of the parcel left after easements, wetlands, floodplain and steep ground.",
    "dischargeable_envelope_acres": "Acres outside every 150-yard safety zone around an occupied building.",
    "dischargeable_envelope_longest_dim_yards": "Longest straight line inside that envelope: a ribbon is useless, a square works.",
    "dwellings_with_line_of_sight": "Neighbouring households that can see into the envelope over the terrain.",
    "adjacent_residential_zoning_pct": "Share of adjoining land zoned residential: the threat is what may be built.",
    "adjacent_planned_sewer": "Adjoining land in a planned sewer service area — the strongest predictor of subdivision.",
    "approved_unbuilt_units_within_2mi": "Homes approved but not yet built within two miles (Frederick publishes these).",
    "adjacent_permanently_eased_acres": "Adjoining farmland under permanent easement: neighbours who stay neighbours.",
    "mprp_tier": "Exposure to the studied 500 kV transmission line, worst first (1, then 2, then 3, then clear).",
    "hv_line_nearest_ft": "Distance to the nearest existing high-voltage corridor.",
    "landlocked_apparent": "No direct road contact found — flagged, never deleted; a deeded easement would not show on a map.",
    "frontage_blocked_by_foreign_parcel": "Someone else's land sits between the parcel and the road.",
    "est_per_acre": "Estimated price per acre from recent arms-length agricultural sales.",
    "corridor_durability_score": "How likely the commute corridor is to hold up: approved homes, traffic growth, programmed widenings.",
    "commute_bwi_peak_min": "Minutes to BWI at a 07:00 departure.",
    "commute_langley_peak_min": "Minutes to Langley at a 07:00 departure.",
    "commute_nova_peak_min": "Minutes to Tysons at a 07:00 departure.",
}
for i, m in enumerate(METRICS, start=2):
    w = float(cfg.shortlist.weights.get(m, 0.0))
    a = wsw.cell(i, 1, m.replace("_", " ")); a.font = Font(name=F, size=10, color=INK)
    b = wsw.cell(i, 2, w); b.font = Font(name=F, size=10, bold=True, color=BLUE)
    b.fill = INPUT; b.number_format = "0.00"; b.border = BOX
    c = wsw.cell(i, 3, "more is better" if w > 0 else "less is better" if w < 0 else "reported only")
    c.font = Font(name=F, size=10, color="66756C")
    d = wsw.cell(i, 4, MEANING.get(m, "")); d.font = Font(name=F, size=10, color=INK)
    d.alignment = Alignment(wrap_text=True, vertical="top")
    a.border = BOX; c.border = BOX; d.border = BOX
    wsw.row_dimensions[i].height = 26
last_w = 1 + len(METRICS)
note = wsw.cell(last_w + 2, 1, "Blue figures on yellow are yours to change. Commute weights ship at 0 by design: the spec reports commute, never filters on it.")
note.font = Font(name=F, size=9, italic=True, color="66756C")
wsw.merge_cells(start_row=last_w + 2, start_column=1, end_row=last_w + 2, end_column=4)
wsw.freeze_panes = "A2"

# ---------------------------------------------------------------- Candidates
wsc = wb.create_sheet("Candidates")
COLS = [
    ("Rank", 7, None), ("Score", 8, "0.000"), ("Score as shipped", 9, "0.000"), ("Shipped rank", 8, "0"),
    ("Account", 13, "@"), ("County", 11, None),
    ("Gross ac", 9, "#,##0.0"), ("Usable ac", 9, "#,##0.0"),
    ("Reachable block ac", 11, "#,##0.0"), ("With a crossing ac", 11, "#,##0.0"),
    ("Envelope ac", 10, "#,##0.0"), ("Envelope length yd", 11, "#,##0"),
    ("Houses in sight", 9, "#,##0"), ("Nearest house yd", 10, "#,##0"),
    ("Backstop ac", 9, "#,##0.0"),
    ("MPRP tier", 8, "0"), ("Adjoining residential ac", 12, "#,##0.0"),
    ("Adjoining eased ac", 11, "#,##0.0"), ("Planned sewer next door", 12, None),
    ("Approved homes 2 mi", 12, "#,##0"),
    ("Est. value", 13, '$#,##0;($#,##0);-'), ("Est. $/ac", 10, '$#,##0;($#,##0);-'),
    ("BWI min", 8, "#,##0"), ("Langley min", 9, "#,##0"), ("Tysons min", 9, "#,##0"),
    ("Ways out", 12, None), ("Corridor durability", 10, "#,##0.0"),
    ("Owner mailbox", 34, None), ("SDAT record", 11, None),
    ("Status", 14, None), ("Notes", 40, None),
]
for i, (h, w, _) in enumerate(COLS, start=1):
    wsc.cell(1, i, h)
    wsc.column_dimensions[get_column_letter(i)].width = w
n0 = len(COLS) + 1
for j, m in enumerate(METRICS):
    c = wsc.cell(1, n0 + j, m)
    c.font = Font(name=F, size=8, color="FFFFFF")
    wsc.column_dimensions[get_column_letter(n0 + j)].width = 10
style_header(wsc, 1, len(COLS))
for j in range(len(METRICS)):
    cell = wsc.cell(1, n0 + j)
    cell.fill = PatternFill("solid", fgColor="8A978F")
    cell.alignment = Alignment(vertical="center", wrap_text=True)

STATUS_COL = [h for h, _, _ in COLS].index("Status") + 1
NOTES_COL = [h for h, _, _ in COLS].index("Notes") + 1
order = list(E.sort_values("shortlist_score", ascending=False).index)
nrows = len(order)


def val(acct, col, default=None):
    if col in P.columns and acct in P.index:
        v = P.at[acct, col]
        if pd.isna(v) if not isinstance(v, (list, dict)) else False:
            return default
        return v
    return default


ways = {"redundant": "two or more", "single_egress": "one road out", "no_route": "no state road found"}
for i, acct in enumerate(order):
    r = i + 2
    score_cell = f"B{r}"
    nrange = f"{get_column_letter(n0)}{r}:{get_column_letter(n0 + len(METRICS) - 1)}{r}"
    wrange = f"Weights!$B$2:$B${last_w}"
    row = [
        f"=RANK({score_cell},$B$2:$B${nrows + 1})",
        f"=SUMPRODUCT({nrange},{wrange})",
        float(E.at[acct, "shortlist_score"]),
        shipped.get(acct),
        acct, val(acct, "county"),
        val(acct, "gross_acres"), val(acct, "usable_acres"),
        val(acct, "largest_contiguous_reachable_acres"), val(acct, "largest_reachable_if_crossings_permitted_acres"),
        val(acct, "dischargeable_envelope_acres"), val(acct, "dischargeable_envelope_longest_dim_yards"),
        val(acct, "dwellings_with_line_of_sight"), val(acct, "nearest_dwelling_yards"),
        val(acct, "candidate_backstop_acres"),
        val(acct, "mprp_tier"), val(acct, "adjacent_residential_zoning_acres"),
        val(acct, "adjacent_permanently_eased_acres"),
        {True: "yes", False: "no"}.get(val(acct, "adjacent_planned_sewer"), "not published"),
        val(acct, "approved_unbuilt_units_within_2mi"),
        val(acct, "est_market_value"), val(acct, "est_per_acre"),
        val(acct, "commute_bwi_peak_min"), val(acct, "commute_langley_peak_min"), val(acct, "commute_nova_peak_min"),
        ways.get(val(acct, "route_redundancy"), ""), val(acct, "corridor_durability_score"),
        val(acct, "owner_mailing_address"), val(acct, "sdat_url"),
        None, None,
    ]
    for j, v in enumerate(row, start=1):
        c = wsc.cell(r, j)
        if isinstance(v, str) and v.startswith("http"):
            c.value = "open"
            c.hyperlink = v
            c.font = Font(name=F, size=10, color="1F6F8B", underline="single")
        else:
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                c.value = float(v)
            elif v is not None:
                c.value = str(v) if not isinstance(v, str) else v
            c.font = Font(name=F, size=10, color=INK)
        fmt = COLS[j - 1][2]
        if fmt:
            c.number_format = fmt
        c.border = BOX
    for j in (STATUS_COL, NOTES_COL):       # Status, Notes: yours to fill
        c = wsc.cell(r, j); c.fill = INPUT; c.font = Font(name=F, size=10, color=BLUE); c.border = BOX
    for j, m in enumerate(METRICS):
        v = N.at[acct, f"score_{m}"] if acct in N.index else None
        c = wsc.cell(r, n0 + j, None if v is None or pd.isna(v) else float(v))
        c.number_format = "0.000"; c.font = Font(name=F, size=9, color="8A978F")
    if shipped.get(acct):
        for j in range(1, 7):
            wsc.cell(r, j).fill = BAND

wsc.freeze_panes = "E2"
wsc.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{nrows + 1}"
wsc.conditional_formatting.add(
    f"B2:B{nrows + 1}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="C7E9C0"))
dv = DataValidation(type="list", formula1='"to look at,called the county,drive by,offer,rejected"', allow_blank=True)
wsc.add_data_validation(dv)
dv.add(f"{get_column_letter(STATUS_COL)}2:{get_column_letter(STATUS_COL)}{nrows + 1}")
wsc.column_dimensions.group(get_column_letter(n0), get_column_letter(n0 + len(METRICS) - 1), hidden=True)

# ---------------------------------------------------------------- Excluded
wse = wb.create_sheet("Excluded")
REASON = {
    "largest_reachable_block_below_acreage_min_even_with_crossings":
        "No 40-acre block you can drive to, even if a stream crossing were permitted",
    "mprp_tier_1_intersects_studied_route": "Intersects a studied route of the 500 kV transmission line",
    "owner_type_government": "Government-owned: not an acquisition candidate",
}
for i, h in enumerate(["Account", "County", "Gross ac", "Why it was excluded"], start=1):
    wse.cell(1, i, h)
style_header(wse, 1, 4)
widths(wse, {"A": 14, "B": 12, "C": 10, "D": 74})
ex = excluded.copy()
ex["order"] = ex["exclusion_reason"].map({k: i for i, k in enumerate(REASON)}).fillna(9)
for i, (_, e) in enumerate(ex.sort_values(["order", "gross_acres"], ascending=[True, False]).iterrows(), start=2):
    for j, v in enumerate([e["account_id"], e.get("county"), e.get("gross_acres"),
                           REASON.get(e["exclusion_reason"], e["exclusion_reason"])], start=1):
        c = wse.cell(i, j, float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else v)
        c.font = Font(name=F, size=10, color=INK); c.border = BOX
        if j == 3:
            c.number_format = "#,##0.0"
wse.freeze_panes = "A2"
wse.auto_filter.ref = f"A1:D{len(ex) + 1}"

# ---------------------------------------------------------------- Owners
wso = wb.create_sheet("Owners")
OW = pd.read_csv("outputs/owner_list.csv")
head = ["Owner mailbox", "Type", "Parcels", "Total ac", "Largest reachable ac", "Best rank", "On the shortlist", "Counties", "Accounts"]
for i, h in enumerate(head, start=1):
    wso.cell(1, i, h)
style_header(wso, 1, len(head))
widths(wso, {"A": 40, "B": 12, "C": 8, "D": 11, "E": 12, "F": 9, "G": 13, "H": 22, "I": 46})
for i, (_, o) in enumerate(OW.iterrows(), start=2):
    vals = [o.get("owner_mailing_address") or o.get("owner_key"), o.get("owner_type"), o.get("parcel_count"),
            o.get("gross_acres_total"), o.get("largest_reachable_acres_max"),
            None if pd.isna(o.get("best_shortlist_rank")) else o.get("best_shortlist_rank"),
            "yes" if o.get("on_shortlist") else "", o.get("counties"), o.get("account_ids")]
    for j, v in enumerate(vals, start=1):
        c = wso.cell(i, j)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            c.value = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else str(v)
        c.font = Font(name=F, size=10, color=INK); c.border = BOX
        if j in (4, 5):
            c.number_format = "#,##0.0"
wso.freeze_panes = "A2"
wso.auto_filter.ref = f"A1:I{len(OW) + 1}"

# ---------------------------------------------------------------- Comps
wsp = wb.create_sheet("Comps")
CO = pd.read_csv("outputs/valuation_comps.csv")
BASIS = {"parcel_area": "share of the sold parcel under easement",
         "sale_point": "sale point only (parcel outside our fabric)",
         "unknown_no_easement_coverage": "unknown — no easement layer covers that county"}
head = ["County", "Sale date", "Price", "Acres", "Land value", "$ per acre", "Eased", "How eased was decided", "Accounts"]
for i, h in enumerate(head, start=1):
    wsp.cell(1, i, h)
style_header(wsp, 1, len(head))
widths(wsp, {"A": 13, "B": 11, "C": 13, "D": 9, "E": 13, "F": 11, "G": 9, "H": 42, "I": 26})
for i, (_, k) in enumerate(CO.sort_values(["county", "land_price_per_acre"]).iterrows(), start=2):
    eased = k.get("eased")
    vals = [k.get("county"), str(k.get("sale_date"))[:10], k.get("price"), k.get("acres"), k.get("land_price"),
            k.get("land_price_per_acre"),
            "" if pd.isna(eased) else ("yes" if eased in (True, "True") else "no"),
            BASIS.get(k.get("eased_basis"), k.get("eased_basis")), k.get("accounts")]
    for j, v in enumerate(vals, start=1):
        c = wsp.cell(i, j)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            c.value = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else str(v)
        c.font = Font(name=F, size=10, color=INK); c.border = BOX
        if j in (3, 5, 6):
            c.number_format = '$#,##0;($#,##0);-'
        if j == 4:
            c.number_format = "#,##0.0"
wsp.freeze_panes = "A2"
wsp.auto_filter.ref = f"A1:I{len(CO) + 1}"
band_note = wsp.cell(len(CO) + 3, 1, "Land value = consideration less the assessed improvement value at sale. "
                                     "A sale whose improvements exceed 80% of the price is a farmstead, not a land comp, and was dropped. "
                                     "Source: MD iMAP MD_PropertySales, arms-length conveyances (codes 1-3) of the last three years.")
band_note.font = Font(name=F, size=9, italic=True, color="66756C")
wsp.merge_cells(start_row=len(CO) + 3, start_column=1, end_row=len(CO) + 3, end_column=9)

wb.save(OUT)
print("wrote", OUT, "| candidates", nrows, "| excluded", len(ex), "| owners", len(OW), "| comps", len(CO))
